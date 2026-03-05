import json
import posixpath
import shutil
from pathlib import Path
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import load_workbook


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _resolve_path(base_file: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_file), target))


def _find_rel_target(rels_root: ET.Element, rel_id: str) -> str:
    for rel in rels_root:
        if rel.attrib.get("Id") == rel_id:
            return rel.attrib.get("Target", "")
    return ""


def extract_sheet_images(source_file: Path, sheet_name: str, output_dir: Path) -> dict[int, str]:
    row_to_image: dict[int, str] = {}

    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(source_file) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        sheet_rid = ""
        for sheet in workbook_root.findall(f".//{{{SHEET_NS}}}sheet"):
            if sheet.attrib.get("name") == sheet_name:
                sheet_rid = sheet.attrib.get(f"{{{DOC_REL_NS}}}id", "")
                break

        if not sheet_rid:
            return row_to_image

        worksheet_target = _find_rel_target(workbook_rels_root, sheet_rid)
        if not worksheet_target:
            return row_to_image
        worksheet_path = _resolve_path("xl/workbook.xml", worksheet_target)

        worksheet_root = ET.fromstring(archive.read(worksheet_path))
        drawing_el = worksheet_root.find(f"{{{SHEET_NS}}}drawing")
        if drawing_el is None:
            return row_to_image

        drawing_rid = drawing_el.attrib.get(f"{{{DOC_REL_NS}}}id", "")
        if not drawing_rid:
            return row_to_image

        worksheet_rels_path = (
            f"{posixpath.dirname(worksheet_path)}/_rels/{posixpath.basename(worksheet_path)}.rels"
        )
        worksheet_rels_root = ET.fromstring(archive.read(worksheet_rels_path))
        drawing_target = _find_rel_target(worksheet_rels_root, drawing_rid)
        if not drawing_target:
            return row_to_image
        drawing_path = _resolve_path(worksheet_path, drawing_target)

        drawing_rels_path = (
            f"{posixpath.dirname(drawing_path)}/_rels/{posixpath.basename(drawing_path)}.rels"
        )
        drawing_rels_root = ET.fromstring(archive.read(drawing_rels_path))

        rid_to_media: dict[str, str] = {}
        for rel in drawing_rels_root.findall(f"{{{REL_NS}}}Relationship"):
            rel_id = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if rel_id and target:
                media_path = _resolve_path(drawing_path, target)
                rid_to_media[rel_id] = media_path

        drawing_root = ET.fromstring(archive.read(drawing_path))
        ns = {
            "xdr": DRAWING_NS,
            "a": A_NS,
        }

        candidates: dict[int, list[tuple[int, str]]] = {}
        anchors = drawing_root.findall("xdr:twoCellAnchor", ns) + drawing_root.findall(
            "xdr:oneCellAnchor", ns
        )
        for anchor in anchors:
            from_el = anchor.find("xdr:from", ns)
            if from_el is None:
                continue

            row_el = from_el.find("xdr:row", ns)
            col_el = from_el.find("xdr:col", ns)
            if row_el is None or col_el is None:
                continue

            blip_el = anchor.find(".//a:blip", ns)
            if blip_el is None:
                continue

            rid = blip_el.attrib.get(f"{{{DOC_REL_NS}}}embed", "")
            media_path = rid_to_media.get(rid)
            if not media_path:
                continue

            row_text = row_el.text
            col_text = col_el.text
            if row_text is None or col_text is None:
                continue

            row_num = int(row_text) + 1
            col_num = int(col_text) + 1
            candidates.setdefault(row_num, []).append((col_num, media_path))

        selected_media = {
            row_num: sorted(media_list, key=lambda item: item[0])[0][1]
            for row_num, media_list in candidates.items()
        }

        media_to_relative: dict[str, str] = {}
        for media_path in sorted(set(selected_media.values())):
            filename = posixpath.basename(media_path)
            output_path = output_dir / filename
            output_path.write_bytes(archive.read(media_path))
            media_to_relative[media_path] = f"data/images/{filename}"

        for row_num, media_path in selected_media.items():
            row_to_image[row_num] = media_to_relative.get(media_path, "")

    return row_to_image


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    source_file = base_dir / "export.xlsx"
    output_dir = base_dir / "data"
    output_file = output_dir / "products.json"

    workbook = load_workbook(source_file, data_only=True)
    sheet = workbook["data"]
    row_to_image = extract_sheet_images(source_file, "data", output_dir / "images")

    products = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        excel_row = idx + 1
        if not any(row):
            continue

        cell_image = normalize(row[5])
        image = row_to_image.get(excel_row, cell_image)

        products.append(
            {
                "id": idx,
                "customer": normalize(row[0]),
                "category": normalize(row[1]),
                "code": normalize(row[2]),
                "name": normalize(row[3]),
                "price": normalize(row[4]),
                "image": image,
                "note": normalize(row[6]),
                "link": normalize(row[7]),
            }
        )

    output_dir.mkdir(exist_ok=True)
    output_file.write_text(json.dumps(products, ensure_ascii=False), encoding="utf-8")

    print(
        f"Exported {len(products)} products to {output_file} "
        f"(images matched: {sum(1 for p in products if p['image'])})"
    )


if __name__ == "__main__":
    main()
